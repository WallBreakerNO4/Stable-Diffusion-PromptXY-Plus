import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image
import ast
from tqdm import tqdm


def process_image(image_path):
    """处理图片，保持原始尺寸和质量"""
    with Image.open(image_path) as img:
        # 转换为RGB模式（如果需要）
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
        
        # 保存为临时PNG文件，使用最高质量设置
        temp_path = image_path.replace('.webp', '_temp.png')
        img.save(temp_path, 'PNG', quality=100, optimize=True)
        return temp_path


def cleanup_temp_files(temp_files):
    """清理临时文件"""
    for temp_file in temp_files:
        try:
            os.remove(temp_file)
        except Exception as e:
            print(f"清理临时文件时出错: {e}")


# 读取图片路径CSV文件
df = pd.read_csv("compressed_image_paths.csv", index_col=0)

# 创建一个新的Excel工作簿
wb = Workbook()
ws = wb.active
ws.title = "压缩图片展示"

# 设置列宽和行高（适应原始图片尺寸）
ws.column_dimensions['A'].width = 35  # 艺术家列
ws.column_dimensions['B'].width = 35  # 提示词列

# 添加标题行
ws['A1'] = "艺术家"
ws['B1'] = "提示词"
ws['C1'] = "生成图片"

# 用于跟踪需要清理的临时文件
temp_files = []

# 创建进度条
total_rows = sum(len(df.columns) for _ in df.index)
progress_bar = tqdm(total=total_rows, desc="生成Excel报告进度")

# 当前行号（从2开始，因为1是标题行）
current_row = 2

# 遍历DataFrame中的每个单元格
for prompt in df.columns:
    for artist in df.index:
        # 获取图片路径列表（存储为字符串，需要解析）
        try:
            image_paths = ast.literal_eval(df.loc[artist, prompt])
            if not isinstance(image_paths, list):
                image_paths = [image_paths]
        except:
            image_paths = []

        # 写入艺术家和提示词
        ws[f'A{current_row}'] = artist
        ws[f'B{current_row}'] = prompt

        # 处理图片
        for idx, img_path in enumerate(image_paths):
            if os.path.exists(img_path):
                try:
                    # 处理图片并获取临时文件路径
                    temp_path = process_image(img_path)
                    temp_files.append(temp_path)

                    # 在Excel中插入图片
                    img = XLImage(temp_path)
                    cell = ws.cell(row=current_row, column=3 + idx)
                    
                    # 设置更大的列宽和行高以适应原始图片
                    ws.column_dimensions[get_column_letter(3 + idx)].width = 60
                    ws.row_dimensions[current_row].height = 400
                    
                    # 添加图片并设置偏移以确保居中显示
                    ws.add_image(img, cell.coordinate)
                except Exception as e:
                    print(f"处理图片时出错 {img_path}: {e}")

        current_row += 1
        progress_bar.update(1)

progress_bar.close()

# 保存Excel文件
wb.save("compressed_images_report.xlsx")

# 清理临时文件
cleanup_temp_files(temp_files)

print("原始尺寸Excel报告已生成: compressed_images_report.xlsx")