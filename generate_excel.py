import ast
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment

# Load the DataFrame from the transposed CSV file
df = pd.read_csv('image_paths.csv', index_col=0)

# Create a new Excel workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Set the row height and column width to fit the images
row_height = (1536 // 3.25) + 10  # Increase the row height more significantly
col_width = (1024 // 18.75) + 10  # Column width remains the same
image_scale = 0.4  # Resize the image to 40% of its original size

# Write the DataFrame to the Excel worksheet
for j, artist_string in enumerate(df.columns, start=2):
    cell = ws.cell(row=1, column=j, value=artist_string)  # Write the artist strings to the first row
    cell.alignment = Alignment(wrapText=True)  # Enable text wrapping
    ws.column_dimensions[chr(64 + j)].width = col_width

ws.column_dimensions['A'].width = col_width // 2  # Set the width of the first column to twice the original width

for i, (prompt_string, row) in enumerate(df.iterrows(), start=2):
    cell = ws.cell(row=i, column=1, value=prompt_string)  # Write the prompt strings to the first column
    cell.alignment = Alignment(wrapText=True)  # Enable text wrapping
    ws.row_dimensions[i].height = row_height
    for j, artist_string in enumerate(df.columns, start=2):
        image_paths = row[artist_string]
        if pd.notna(image_paths):
            for image_path in ast.literal_eval(image_paths):  # Use ast.literal_eval to convert string back to list
                img = ExcelImage(image_path)
                img.width = img.width * image_scale
                img.height = img.height * image_scale
                img.anchor = ws.cell(row=i, column=j).coordinate
                ws.add_image(img)

# Save the workbook
wb.save('output_images.xlsx')