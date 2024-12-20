import csv
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from lib.convert_braces import convert_braces
from lib.construct_json_payload import construct_json_payload_with_artist
from lib.fetch_and_return_images import fetch_and_return_images
import hashlib


# Function to generate a short hash for a given string
def generate_short_hash(input_string):
    return hashlib.md5(input_string.encode()).hexdigest()[:10]


# Ensure the images directory exists
os.makedirs('../images', exist_ok=True)

# Read CSV files
with (open('../prompts/artist_strings.csv', 'r', encoding='utf-8-sig') as artist_strings_csv,
      open('../prompts/prompt_string.csv', 'r', encoding='utf-8-sig') as prompt_string_csv):
    artist_string_reader = csv.reader(artist_strings_csv)
    prompt_string_reader = csv.reader(prompt_string_csv)
    artist_strings = []
    prompt_strings = []
    api_url = "http://192.168.20.100:7860/sdapi/v1/txt2img"

    for row in artist_string_reader:
        artist_strings.append(convert_braces(row[0]))
    for row in prompt_string_reader:
        prompt_strings.append(row[0])

    # Create a DataFrame to store the results
    df = pd.DataFrame(index=prompt_strings, columns=artist_strings)

    for artist_string in artist_strings:
        for prompt_string in prompt_strings:
            json_payload = construct_json_payload_with_artist(artist_string, prompt_string)
            images = fetch_and_return_images(api_url, json_payload)
            image_paths = []
            for idx, image in enumerate(images):
                artist_hash = generate_short_hash(artist_string)
                prompt_hash = generate_short_hash(prompt_string)
                image_path = f"images/{artist_hash}_{prompt_hash}_{idx}.png"
                image.save(image_path)
                image_paths.append(image_path)
                print(f"Saved image {idx + 1} for artist:\"{artist_string}\" and prompt:\"{prompt_string}\"")
            # Store the image paths in the DataFrame
            df.at[prompt_string, artist_string] = image_paths

# Create a new Excel workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Set the row height and column width to fit the images
row_height = (1536 // 1.3) + 10  # Increase the row height more significantly
col_width = (1024 // 7.5) + 10  # Column width remains the same

# Write the DataFrame to the Excel worksheet in batches
batch_size = 100  # Define the batch size
for i in range(0, len(df.index), batch_size):
    batch_df = df.iloc[i:i + batch_size]
    for j, artist_string in enumerate(batch_df.columns, start=2):
        ws.cell(row=1, column=j, value=artist_string)
        ws.column_dimensions[chr(64 + j)].width = col_width
    for i, prompt_string in enumerate(batch_df.index, start=2):
        ws.cell(row=i, column=1, value=prompt_string)
        ws.row_dimensions[i].height = row_height
        for j, artist_string in enumerate(batch_df.columns, start=2):
            image_paths = batch_df.at[prompt_string, artist_string]
            if image_paths:
                for image_path in image_paths:
                    img = ExcelImage(image_path)
                    img.anchor = ws.cell(row=i, column=j).coordinate
                    ws.add_image(img)

# Save the workbook
wb.save('output_images.xlsx')
