import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage

# Ensure the images directory exists
images_dir = '../images'
if not os.path.exists(images_dir):
    raise FileNotFoundError(f"The directory '{images_dir}' does not exist.")

# Create a DataFrame to store the image paths
image_files = [f for f in os.listdir(images_dir) if f.endswith('.png')]
df = pd.DataFrame(columns=['Image Path'])

# Collect image paths in a list of dictionaries
image_data = [{'Image Path': os.path.join(images_dir, image_file)} for image_file in image_files]

# Concatenate the list of dictionaries into the DataFrame
df = pd.concat([df, pd.DataFrame(image_data)], ignore_index=True)

# Create a new Excel workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Set the row height and column width to fit the images
row_height = (1536 // 1.3) + 10  # Increase the row height more significantly
col_width = (1024 // 7.5) + 10  # Column width remains the same

# Write the DataFrame to the Excel worksheet
for i, row in df.iterrows():
    ws.cell(row=i + 2, column=1, value=row['Image Path'])
    ws.row_dimensions[i + 2].height = row_height
    img = ExcelImage(row['Image Path'])
    img.anchor = ws.cell(row=i + 2, column=2).coordinate
    ws.add_image(img)

# Set the column width for the image column
ws.column_dimensions['A'].width = col_width

# Save the workbook
wb.save('output_images.xlsx')
